

import time
import os
import gc
import json
from openpyxl import load_workbook
import os

def clean_header(value):
    if value is None:
        return ""
    value = str(value).strip().lower()
    if value in ("", "none", "null", "nan"):
        return ""
    return value

def log_message(log_file, message):
    os.makedirs(os.path.dirname(log_file), exist_ok=True)
    with open(log_file, 'a', encoding='utf-8') as log:
        log.write(message + '\n')
    print(message)

def insert_columns_from_main(main_file, raw_file, current_department, log_file):
    area_column = None
    rf_column = None
    upload_direct = False
    output_value = None
    area_already_inserted = False
    
    # Load MAIN file
    main_wb = load_workbook(main_file)
    main_ws = main_wb.active

    # Load RAW file
    raw_wb = load_workbook(raw_file)
    raw_ws = raw_wb.active
    
    print("Checking first 3 rows for 'Fulfillment Center'...")
    log_message(log_file, "Checking first 3 rows for 'Fulfillment Center'...")
    fulfillment_found = False
    for row in range(1, 4):  # rows 1 to 3
        for col in range(1, raw_ws.max_column + 1):
            cell_value = raw_ws.cell(row=row, column=col).value
            if cell_value and "fulfillment center" in str(cell_value).lower():
                fulfillment_found = True
                break
        if fulfillment_found:
            break

    if fulfillment_found:
        print("Found 'Fulfillment Center' -> Deleting first 3 rows...")
        log_message(log_file, "Found 'Fulfillment Center' -> Deleting first 3 rows...")
        raw_ws.delete_rows(1, 3)
        raw_wb.save(raw_file)
        print("Deleted first 3 rows and saved file.\n")
        log_message(log_file, "Deleted first 3 rows and saved file.\n")
    else:
        print("No 'Fulfillment Center' found → No rows deleted.\n")
        log_message(log_file, "No Fulfillment Center found → No rows deleted.\n")
    
    for row in range(2, raw_ws.max_row + 1):
        cell = raw_ws.cell(row=row, column=2)  # Column B
        value = cell.value

        if value is None:
            continue
        cleaned = str(value).replace("*", "").strip()
        # If the cleaned value is a valid integer (including negative), convert it
        if cleaned.lstrip("-").isdigit():
            cell.value = int(cleaned)
        else:
            cell.value = cleaned

    print("Column B cleanup completed successfully.")
        
        
    # Read headers from Main
    headers_main = [clean_header(c.value) for c in main_ws[1]]

    # Find column index of required fields
    idx_department       = headers_main.index("department") + 1
    idx_subdept          = headers_main.index("subdepartment") + 1
    idx_subdept_name     = headers_main.index("subdepartmentname") + 1
    idx_lookup_table     = headers_main.index("lookuptablename") + 1
    idx_insert_amount    = headers_main.index("insertcolumnsamount") + 1
    idx_insert_position  = headers_main.index("insertcolumnposition") + 1
    idx_column_names     = headers_main.index("columnnames") + 1
    idx_outputname       = headers_main.index("outputname") + 1

    print("\n=== PROCESSING MAIN FILE ROWS ===")
    log_message(log_file, "\n=== PROCESSING MAIN FILE ROWS ===")
    target_col = 6
    existing_header = raw_ws.cell(row=1, column=target_col).value
    existing_header_clean = clean_header(existing_header)
    print(f"Existing header at column F: {existing_header}")
    if existing_header_clean != "rf" and existing_header_clean != "area":
        print("RF column not found in column F, inserting Area column.")
        log_message(log_file, "RF column not found in column F, inserting Area column.")
        raw_ws.insert_cols(target_col)
        raw_ws.cell(row=1, column=target_col).value = "Area"
        area_column = "created"
        print("Area column added in column F")
        log_message(log_file, "Area column added in column F")
    else:
        print("Area/RF column already exists in column F")
        log_message(log_file, "Area/RF column already exists in column F")

    inserted_keys = set()
    for row in range(2, main_ws.max_row + 1):

        department         = clean_header(main_ws.cell(row=row, column=idx_department).value)
        lookup_table       = clean_header(main_ws.cell(row=row, column=idx_lookup_table).value)
        output_name        = clean_header(main_ws.cell(row=row, column=idx_outputname).value)

        raw_headers = [str(c.value).strip().lower() if c.value else "" for c in raw_ws[1]]
        if department != current_department:
            continue
        if lookup_table == "":
            upload_direct = True
            for r in range(2, raw_ws.max_row + 1):
                raw_ws.cell(row=r, column=target_col).value = output_name.upper()
            continue  
        # else:
        # Read Insert instructions
        amount = main_ws.cell(row=row, column=idx_insert_amount).value
        position = main_ws.cell(row=row, column=idx_insert_position).value
        col_name = main_ws.cell(row=row, column=idx_column_names).value

        if not amount or amount <= 0:
            continue
        key = f"{position}_{col_name}"        # unique key per column & position
        # Read RAW file headers
        # raw_headers = [str(c.value).strip().lower() if c.value else "" for c in raw_ws[1]]

        existing_header = raw_ws.cell(row=1, column=position).value
        print(f"Existing header at position {position}: {existing_header}")
        log_message(log_file, f"Existing header at position {position}: {existing_header}")
        if existing_header and existing_header.strip().upper() == "RF":
            print(f"SKIPPED inserting '{col_name}' at position {position} because 'RF' column already exists.")
            log_message(log_file,f"SKIPPED inserting '{col_name}' at position {position} because 'RF' column already exists.")
            continue
        
        if col_name.strip().lower() == "area" and "area" in [header.lower() for header in raw_headers]:
            print(f"SKIPPED inserting 'Area' because Area column already exists.")
            log_message(log_file, f"SKIPPED inserting 'Area' because Area column already exists.")
            continue

            # Skip if already inserted
        if key in inserted_keys:
            continue

        inserted_keys.add(key)

        
        raw_ws.insert_cols(position, amount)

        for i in range(amount):
            raw_ws.cell(row=1, column=position + i).value = col_name
        if col_name.lower() == "area":
            area_column = "created"

        if col_name.lower() == "rf":
            rf_column = "created"
        print(f"Inserting {amount} columns at position {position} named '{col_name}' into RAW file")
        log_message(log_file, f"Inserting {amount} columns at position {position} named '{col_name}' into RAW file")
    

    raw_wb.save(raw_file)
    print(f"Updated RAW file saved: {raw_file}")
    log_message(log_file, f"Updated RAW file saved: {raw_file}")

    del main_wb
    del raw_wb
    # gc.collect()
    time.sleep(2)
    return area_column, rf_column, upload_direct


