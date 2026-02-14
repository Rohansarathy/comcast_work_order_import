
import time
import os
import gc
import json
from openpyxl import load_workbook
from selenium.webdriver.support import expected_conditions as EC

credentials_file = 'accpet_loadjob.json'
with open(credentials_file, 'r') as file:
    credentials = json.load(file)


# raw_excel = r"C:\Users\RohansarathyGoudhama\OneDrive - Yitro Technology Solutions Pvt Ltd\Documents\bigs.xlsx"
# folder_path = credentials['Main_file']
# wb = load_workbook(raw_excel)
# ws = wb.active 





# def insert_column_amount(folder_path, raw_excel):

#     wb = load_workbook(folder_path)
#     ws = wb.active

#     # Read header row
#     headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]

#     # Get column indexes (1-based)
#     try:
#         col_amount_idx = headers.index("InsertColumnsAmount") + 1
#         col_position_idx = headers.index("InsertColumnPosition") + 1
#         col_name_idx = headers.index("ColumnNames") + 1
#     except ValueError:
#         raise Exception("InsertColumnsAmount, InsertColumnPosition, or ColumnNames headers are missing.")

#     # Loop rows & process insertions
#     for row in range(2, ws.max_row + 1):

#         amount = ws.cell(row=row, column=col_amount_idx).value
#         position = ws.cell(row=row, column=col_position_idx).value
#         col_name = ws.cell(row=row, column=col_name_idx).value

#         if amount and amount > 0:
#             print(f"Row {row}: Adding {amount} column(s) at position {position} named '{col_name}'")

#             # Insert new columns
#             ws.insert_cols(position, amount)

#             # Update header name for each inserted column
#             for i in range(amount):
#                 ws.cell(row=1, column=position + i).value = col_name

#     # Save updated file
#     wb.save(raw_excel )
#     print(f"Updated file saved: {raw_excel }")

#     del wb
#     gc.collect()
#     time.sleep(2)


# # ---- Call the function here ----
# insert_column_amount(folder_path, raw_excel )
def insert_columns_from_main(folder_path, raw_file):
    area_column = None
    rf_column = None
    # Load MAIN file (where instructions exist)
    main_wb = load_workbook(folder_path)
    main_ws = main_wb.active

    # Load RAW file (where columns should be inserted)
    raw_wb = load_workbook(raw_file)
    raw_ws = raw_wb.active

    # Read headers from MAIN file
    headers = [str(c.value).strip() if c.value else "" for c in main_ws[1]]
    

    col_amount_idx = headers.index("InsertColumnsAmount") + 1
    col_position_idx = headers.index("InsertColumnPosition") + 1
    col_name_idx = headers.index("ColumnNames") + 1

    # TRACK already inserted columns
    inserted_keys = set()

    # Loop MAIN rows and insert into RAW file
    for row in range(2, main_ws.max_row + 1):

        amount = main_ws.cell(row=row, column=col_amount_idx).value
        position = main_ws.cell(row=row, column=col_position_idx).value
        col_name = main_ws.cell(row=row, column=col_name_idx).value

        # if amount and amount > 0:
        #     key = f"{position}_{col_name}"  
        if not amount or amount <= 0:
            continue
        key = f"{position}_{col_name}"        # unique key per column & position
        # Read RAW file headers
        raw_headers = [str(c.value).strip().lower() if c.value else "" for c in raw_ws[1]]

        # Skip if RF exists
        # if col_name.lower() == "rf" and "rf" in raw_headers:
        #     print("SKIPPING: 'RF' column already exists. No new RF column added.")
        #     continue

        # Skip if AREA exists
        # if col_name.lower() == "area" and "area" in raw_headers:
        #     print("SKIPPING: 'Area' column already exists.")
        #     continue
        existing_header = raw_ws.cell(row=1, column=position).value
        print(f"Existing header at position {position}: {existing_header}")
        if existing_header and existing_header.strip().upper() == "RF":
            print(f"SKIPPED inserting '{col_name}' at position {position} because 'RF' column already exists.")
            continue
        
        if col_name.strip().lower() == "area" and "area" in [header.lower() for header in raw_headers]:
            print(f"SKIPPED inserting 'Area' because Area column already exists.")
            continue

            # Skip if already inserted
        if key in inserted_keys:
            continue

        inserted_keys.add(key)

        
        raw_ws.insert_cols(position, amount)

        for i in range(amount):
            raw_ws.cell(row=1, column=position + i).value = col_name
        if col_name.lower() == "area":
            area_column = "created"

        if col_name.lower() == "rf":
            rf_column = "created"
        print(f"Inserting {amount} column(s) at position {position} named '{col_name}' into RAW file")


    raw_wb.save(raw_file)
    print(f"Updated RAW file saved: {raw_file}")

    del main_wb
    del raw_wb
    gc.collect()
    time.sleep(2)
    return area_column, rf_column
# insert_columns_from_main(folder_path, raw_excel)
















# ws.delete_rows(1,3)
# print("Deleted the first 3 rows.")
# wb.save(raw_excel)
# print(f"Cleaned file saved at: {raw_excel}")

# first row after deletion

# raw_headers = [cell.value for cell in ws[1]]
# headers = [str(h).strip() if h else "" for h in raw_headers]
# headers_lower = [h.lower() for h in headers]

# Area = "area" in headers_lower
# RF = "rf" in headers_lower

# if Area:
#     print("RF header found.")
# elif RF:
#     print("Area header found.")
# else:
#     print("Area and RF not found. Inserting new 'Area' column...")
#     try:
#         sp_col = headers.index("SP") + 1
#         units_col = headers.index("Units") + 1
#     except ValueError:
#         raise Exception("SP or Units header not found. Cannot insert Area column.")
#     insert_position = sp_col + 1
#     ws.insert_cols(insert_position)
#     # Set header name
#     ws.cell(row=1, column=insert_position, value="Area")
#     wb.save(raw_excel)
#     print(f"Cleaned file saved at: {raw_excel}")


    



