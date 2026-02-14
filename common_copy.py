

# import time
# import os
# import gc
# import json
# from openpyxl import load_workbook
# import os

# def clean_header(value):
#     """Normalize header values: remove spaces, convert None/NULL to empty."""
#     if value is None:
#         return ""
#     value = str(value).strip().lower()
#     if value in ("", "none", "null", "nan"):
#         return ""
#     return value


# def insert_columns_from_main(main_file, raw_file):
#     area_column = None
#     rf_column = None
#     upload_direct = False

#     # Load MAIN file
#     main_wb = load_workbook(main_file)
#     main_ws = main_wb.active

#     # Load RAW file
#     raw_wb = load_workbook(raw_file)
#     raw_ws = raw_wb.active

#     # Read headers from Main
#     headers_main = [clean_header(c.value) for c in main_ws[1]]

#     # Find column index of required fields
#     idx_department       = headers_main.index("department") + 1
#     idx_subdept          = headers_main.index("subdepartment") + 1
#     idx_subdept_name     = headers_main.index("subdepartmentname") + 1
#     idx_lookup_table     = headers_main.index("lookuptablename") + 1
#     idx_insert_amount    = headers_main.index("insertcolumnsamount") + 1
#     idx_insert_position  = headers_main.index("insertcolumnposition") + 1
#     idx_column_names     = headers_main.index("columnnames") + 1

#     print("\n=== PROCESSING MAIN FILE ROWS ===")

#     inserted_keys = set()

#     for row in range(2, main_ws.max_row + 1):

#         department         = clean_header(main_ws.cell(row=row, column=idx_department).value)
#         print(f"\nProcessing Department: {department}")
#         sub_department     = clean_header(main_ws.cell(row=row, column=idx_subdept).value)
#         sub_departmentName = clean_header(main_ws.cell(row=row, column=idx_subdept_name).value)
#         lookup_table       = clean_header(main_ws.cell(row=row, column=idx_lookup_table).value)

#         if lookup_table == "":
#             print("LookupTableName is empty → DIRECT UPLOAD REQUIRED")
#             upload_direct = True
#             break
#         else:

#             # if lookup_table == "":
#             #     print("LookupTableName is empty -> uploading file directly to Fuse.")
                
#             #     continue

#             # RULE: If any other required field is empty -> skip row
#             # if (sub_department == "" or sub_departmentName == ""):
#             #     print("SKIPPED — Missing required fields")
#             #     continue

#             # Read Insert instructions
#             amount = main_ws.cell(row=row, column=idx_insert_amount).value
#             position = main_ws.cell(row=row, column=idx_insert_position).value
#             col_name = main_ws.cell(row=row, column=idx_column_names).value

#             # if amount and amount > 0:
#             #     key = f"{position}_{col_name}"  
#             if not amount or amount <= 0:
#                 continue
#             key = f"{position}_{col_name}"        # unique key per column & position
#             # Read RAW file headers
#             raw_headers = [str(c.value).strip().lower() if c.value else "" for c in raw_ws[1]]

#             existing_header = raw_ws.cell(row=1, column=position).value
#             print(f"Existing header at position {position}: {existing_header}")
#             if existing_header and existing_header.strip().upper() == "RF":
#                 print(f"SKIPPED inserting '{col_name}' at position {position} because 'RF' column already exists.")
#                 continue
            
#             if col_name.strip().lower() == "area" and "area" in [header.lower() for header in raw_headers]:
#                 print(f"SKIPPED inserting 'Area' because Area column already exists.")
#                 continue

#                 # Skip if already inserted
#             if key in inserted_keys:
#                 continue

#             inserted_keys.add(key)

            
#             raw_ws.insert_cols(position, amount)

#             for i in range(amount):
#                 raw_ws.cell(row=1, column=position + i).value = col_name
#             if col_name.lower() == "area":
#                 area_column = "created"

#             if col_name.lower() == "rf":
#                 rf_column = "created"
#             print(f"Inserting {amount} column(s) at position {position} named '{col_name}' into RAW file")


#     raw_wb.save(raw_file)
#     print(f"Updated RAW file saved: {raw_file}")

#     # del main_wb
#     # del raw_wb
#     # gc.collect()
#     time.sleep(2)
#     return area_column, rf_column, upload_direct





import time
import os
import gc
import json
from openpyxl import load_workbook
import os

def clean_header(value):
    """Normalize header values: remove spaces, convert None/NULL to empty."""
    if value is None:
        return ""
    value = str(value).strip().lower()
    if value in ("", "none", "null", "nan"):
        return ""
    return value


def insert_columns_from_main(main_file, raw_file, current_department):
    area_column = None
    rf_column = None
    upload_direct = False
    lookup_found = False
    output_value = None

    # Load MAIN file
    main_wb = load_workbook(main_file)
    main_ws = main_wb.active

    # Load RAW file
    raw_wb = load_workbook(raw_file)
    raw_ws = raw_wb.active

    # Read headers from Main
    headers_main = [clean_header(c.value) for c in main_ws[1]]

    # Find column index of required fields
    idx_department       = headers_main.index("department") + 1
    idx_subdept          = headers_main.index("subdepartment") + 1
    idx_subdept_name     = headers_main.index("subdepartmentname") + 1
    idx_lookup_table     = headers_main.index("lookuptablename") + 1
    idx_insert_amount    = headers_main.index("insertcolumnsamount") + 1
    idx_insert_position  = headers_main.index("insertcolumnposition") + 1
    idx_column_names     = headers_main.index("columnnames") + 1
    idx_outputname       = headers_main.index("outputname") + 1

    print("\n=== PROCESSING MAIN FILE ROWS ===")

    inserted_keys = set()

    for row in range(2, main_ws.max_row + 1):

        department         = clean_header(main_ws.cell(row=row, column=idx_department).value)
        print(f"\nProcessing Department: {department}")
        sub_department     = clean_header(main_ws.cell(row=row, column=idx_subdept).value)
        sub_departmentName = clean_header(main_ws.cell(row=row, column=idx_subdept_name).value)
        lookup_table       = clean_header(main_ws.cell(row=row, column=idx_lookup_table).value)
        output_name        = clean_header(main_ws.cell(row=row, column=idx_outputname).value)

        # if department != current_department:
        #     continue
        # lookup_found = True
        # if lookup_table == "":  
        #     print("LookupTableName is empty -> DIRECT UPLOAD REQUIRED")
        #     upload_direct = True
        #     # output_value = output_name.lower()
        #     break
        
        # Read Insert instructions
        amount = main_ws.cell(row=row, column=idx_insert_amount).value
        position = main_ws.cell(row=row, column=idx_insert_position).value
        col_name = main_ws.cell(row=row, column=idx_column_names).value

        if not amount or amount <= 0:
            continue
        key = f"{position}_{col_name}"        # unique key per column & position
        # Read RAW file headers
        raw_headers = [str(c.value).strip().lower() if c.value else "" for c in raw_ws[1]]

        existing_header = raw_ws.cell(row=1, column=position).value
        print(f"Existing header at position {position}: {existing_header}")
        if existing_header and existing_header.strip().upper() == "RF":
            print(f"SKIPPED inserting '{col_name}' at position {position} because 'RF' column already exists.")
            continue
        
        if col_name.strip().lower() == "area" and "area" in [header.lower() for header in raw_headers]:
            print(f"SKIPPED inserting 'Area' because Area column already exists.")
            continue

            # Skip if already inserted
        if key in inserted_keys:
            continue

        inserted_keys.add(key)

        
        raw_ws.insert_cols(position, amount)

        for i in range(amount):
            raw_ws.cell(row=1, column=position + i).value = col_name
        if col_name.lower() == "area":
            area_column = "created"

        if col_name.lower() == "rf":
            rf_column = "created"
        print(f"Inserting {amount} columns at position {position} named '{col_name}' into RAW file")
    

    raw_wb.save(raw_file)
    print(f"Updated RAW file saved: {raw_file}")

    del main_wb
    del raw_wb
    # gc.collect()
    time.sleep(2)
    # if not lookup_found:
    #     upload_direct = False
    return area_column, rf_column, upload_direct