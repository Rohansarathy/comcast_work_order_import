

from openpyxl import load_workbook

BIGS_MAPPING = {
    "bigs1": "Big South Eastern",
    "bigs2": "183960002 - Big South Central",
}

def update_completed_status(folder_path, dept, remark_text):
    if not folder_path.lower().endswith(".xlsx"):
        print(f"Skipping non-.xlsx file: {folder_path}")
        return

    wb = load_workbook(folder_path)
    ws = wb.active

    # Read headers
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]

    # Ensure Remarks column
    if "Remarks" not in headers:
        ws.cell(row=1, column=len(headers) + 1, value="Remarks")
        headers.append("Remarks")

    remarks_col = headers.index("Remarks") + 1

    # Required columns
    try:
        dept_col = headers.index("Department") + 1
        subdept_col = headers.index("SubDepartmentName") + 1
    except ValueError:
        wb.close()
        raise Exception("Required column missing: Department or SubDepartmentName")

    target_subdept = BIGS_MAPPING.get(dept)
    updated = False

    # Update rows
    for r in range(2, ws.max_row + 1):
        dept_val = ws.cell(r, dept_col).value
        subdept_val = ws.cell(r, subdept_col).value

        if dept_val is None:
            continue

        dept_val = str(dept_val).strip().lower()
        subdept_val = str(subdept_val).strip() if subdept_val else ""

        # BIGS special handling
        if dept in BIGS_MAPPING:
            if dept_val == "bigs" and subdept_val == target_subdept:
                ws.cell(row=r, column=remarks_col, value=remark_text)
                updated = True

        # Normal departments
        else:
            if dept_val == dept:
                ws.cell(row=r, column=remarks_col, value=remark_text)
                updated = True

    wb.save(folder_path)
    wb.close()

    if updated:
        print(f"\033[94mUpdated Remarks = {remark_text} for {dept}\033[0m")
    else:
        print(f"\033[91mNo matching row found for {dept}\033[0m")
