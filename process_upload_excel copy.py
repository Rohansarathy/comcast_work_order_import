import os
import json
import time
from selenium import webdriver
from datetime import datetime
from openpyxl import load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support import expected_conditions as EC

from fuselogin import fuse_login
# from raw_excel import insert_columns_from_main
from bulk_upload import upload_raw_files
from edit_columns import insert_columns_from_main

credentials_file = 'accpet_loadjob.json'
with open(credentials_file, 'r') as file:
    credentials = json.load(file)

def log_message(log_file_path, message):
    os.makedirs(os.path.dirname(log_file_path), exist_ok=True)
    with open(log_file_path, 'a', encoding='utf-8') as log:
        log.write(message + '\n')
    print(message)
log_file = f"{credentials['Logfile']}\\work_order_imprt_logs.txt"

chrome_options = Options()
chrome_options.add_argument("--log-level=3")
chrome_options.add_argument(r"--user-data-dir=C:\ChromeSeleniumProfiles\workorder")
chrome_options.add_experimental_option("detach", True)

service = Service(ChromeDriverManager("142.0.7444.60").install())
driver = webdriver.Chrome(service=service, options=chrome_options)
driver.maximize_window()
driver.execute_cdp_cmd("Page.bringToFront", {})

FuseLogin = fuse_login(driver, credentials, log_file)
base_folder  = r"C:\Users\RohansarathyGoudhama\Downloads\work_order_import"
folder_path = credentials['Main_file']
extaction_file = credentials['extaction_file']

# Mapping RAW filename → Lookup sheet name
key_to_lookup = {
    "mia1": "Miami.Broward VLookup",
    "pit1": "Keystone",
    "flor": "FTM V-Lookup",
    "har1": "WNE V-lookup",
    "ne01": "GBR V-Lookup",
    "bigs": "Big South V Lookup",
    "atl1": "Big South V Lookup",
    "chi2": "Big South V Lookup",
    "denv": "Colorado V Lookup",
    "sea1": "Seattle V Lookup",
}

departments = [
    "mia1","jax1","belt","bay1","free","pit1","port","flor",
    "har1","ne01","hou1","bigs","atl1","bigs","denv",
    "salt","sea1","msp1","chi2"
]
today = datetime.today().strftime("%m-%d-%Y") 
# def get_raw_key(raw_filename):
#     base = os.path.basename(raw_filename).split(".")[0]
#     # Split by underscore and take the last part
#     if "_" in base:
#         return base.split("_")[-1].lower()
#     return base.lower()


def process_raw_file(folder_path, raw_excel, date_folder, extaction_file, department):
    print("PROCESSING RAW FILE:", raw_excel)
    # 1) First INSERT REQUIRED COLUMNS into RAW file
    print("STEP 1: Checking & inserting missing columns from MAIN\n")
    area_column, rf_column, upload_direct = insert_columns_from_main(folder_path, raw_excel, department)
    if upload_direct:
        print("Upload Direct Raw files.\n")
        upload_raw_files(driver, date_folder, credentials, raw_excel, department, log_file)
        return
    # else:
    if not (area_column or rf_column):
        print("No new columns inserted. Columns already existed.\n")

    raw_wb = load_workbook(raw_excel)
    raw_ws = raw_wb.active
    print("Checking first 3 rows for 'Fulfillment Center'...")
    fulfillment_found = False
    for row in range(1, 4):  # rows 1 to 3
        for col in range(1, raw_ws.max_column + 1):
            cell_value = raw_ws.cell(row=row, column=col).value
            if cell_value and "fulfillment center" in str(cell_value).lower():
                fulfillment_found = True
                break
        if fulfillment_found:
            break

    if fulfillment_found:
        print("Found 'Fulfillment Center' -> Deleting first 3 rows...")
        raw_ws.delete_rows(1, 3)
        raw_wb.save(raw_excel)
        print("Deleted first 3 rows and saved file.\n")
    else:
        print("No 'Fulfillment Center' found → No rows deleted.\n")
    # department = os.path.basename(os.path.dirname(os.path.dirname(raw_excel)))
    # department = department.lower()

    print(f"Department detected from folder: {department}")

    if department not in key_to_lookup:
        print(f"No lookup mapping found for department '{department}'\n")
        return

    lookup_sheet = key_to_lookup[department]
    print(f"RAW - Lookup sheet selected: {lookup_sheet}")
    raw_wb = load_workbook(raw_excel)
    raw_ws = raw_wb.active
    headers = [str(c.value).strip() if c.value else "" for c in raw_ws[1]]

    if "RteC" not in headers:
        print("RteC column missing in RAW file.\n")
        return

    rtec_col = headers.index("RteC") + 1

    # Read RteC values
    rte_values = []
    for row in range(2, raw_ws.max_row + 1):
        value = raw_ws.cell(row=row, column=rtec_col).value
        if value not in (None, ""):
            rte_values.append(str(value).strip())

    print(f"Total RteC values found: {len(rte_values)}")

    lookup_wb = load_workbook(extaction_file)

    if lookup_sheet not in lookup_wb.sheetnames:
        print(f"Lookup sheet '{lookup_sheet}' not in extraction file.\n")
        return

    lookup_ws = lookup_wb[lookup_sheet]

    # MATCHING
    matched_rows = []
    for row in lookup_ws.iter_rows(min_row=2, values_only=True):
        if row and row[0] and str(row[0]).strip() in rte_values:
            clean = [v for v in row if v is not None]
            matched_rows.append(clean)


    print(f"\nMatched {len(matched_rows)} rows:")
    for r in matched_rows:
        print("--->", r)

    if not matched_rows:
        print("No matching RteC values.\n")
        return

    # Target Column
    headers_lower = [h.lower() for h in headers]

    if "area" in headers_lower:
        target_col = headers_lower.index("area") + 1
    elif "rf" in headers_lower:
        target_col = headers_lower.index("rf") + 1

    else:
        print("Error: AREA/RF column still not found after insertion.\n")
        return

    lookup_dict = {row[0]: row[1] for row in matched_rows}

    update_count = 0
    for row in range(2, raw_ws.max_row + 1):
        rtec_value = raw_ws.cell(row=row, column=rtec_col).value
        if not rtec_value:
            continue

        rtec_value = str(rtec_value).strip()
        if rtec_value in lookup_dict:
            raw_ws.cell(row=row, column=target_col).value = lookup_dict[rtec_value]
            update_count += 1

    raw_wb.save(raw_excel)
    print(f"\n Updated {update_count} rows successfully.")
    print("Uploading file after lookup update...")
    upload_raw_files(driver, date_folder, credentials, raw_excel, department, log_file)
    print("*****************************************************\n")

# process_raw_file(folder_path, raw_excel, extaction_file)
for dept in departments:
    print("\n======================================")
    print(f"Checking Department: {dept}")

    dept_folder = os.path.join(base_folder, dept)
    if not os.path.isdir(dept_folder):
        print(f"Department folder not found: {dept_folder}")
        continue

    date_folder = os.path.join(dept_folder, today)

    if not os.path.isdir(date_folder):
        print(f"Date folder not found: {date_folder}")
        continue

    excel_files = [
        f for f in os.listdir(date_folder)
        if f.lower().endswith(".xlsx")
    ]

    if not excel_files:
        print("No Excel file found in date folder")
        continue

    for excel_name in excel_files:
        excel_path  = os.path.join(date_folder, excel_name)
        print(f"Processing file: {excel_path}")

        process_raw_file(folder_path, excel_path, date_folder, extaction_file, dept)