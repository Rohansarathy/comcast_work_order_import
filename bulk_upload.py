
import os
import time
from datetime import datetime
from Sendmail import Sendmail
from openpyxl import load_workbook
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import TimeoutException

from error_logs import check_error_logs
from excel_utils import update_completed_status

def log_message(log_file, message):
    with open(log_file, 'a', encoding='utf-8') as log:
        log.write(message + '\n')
    print(message)

def ids_with_empty_remarks(folder_path):
    wb = load_workbook(folder_path, read_only=True, data_only=True)
    ws = wb.active
    headers = {str(c.value).strip().lower(): i for i, c in enumerate(ws[1], start=1)}
    dept_column = headers.get("department")
    remarks_col = headers.get("remarks")
    empties = []
    if dept_column:
        for r in range(2, ws.max_row + 1):
            pid = ws.cell(row=r, column=dept_column).value
            rem = ws.cell(row=r, column=remarks_col).value if remarks_col else None
            if pid and (rem is None or str(rem).strip() == ""):
                empties.append(str(pid).strip())
    wb.close()
    return list(empties)

def upload_raw_files(driver, folder_path, date_folder, credentials, raw_excel, department, log_file):
    upload_success = False
    try:
        WebDriverWait(driver, 5).until(EC.visibility_of_element_located((By.XPATH, '//a[text()="Work Order"]')))
        fuse_login = True
    except Exception:
        fuse_login = False
        
    if fuse_login:
        try:
        # process_raw_file(folder_path, raw_excel, extaction_file)
            print("Login successful. Ready to process files.")
            log_message(log_file, "Going to to upload the Excel files")
            time.sleep(5)
            WebDriverWait(driver, 30).until(EC.visibility_of_element_located((By.XPATH, "//a[normalize-space(text())='Work Order']"))).click()
            time.sleep(2)
            WebDriverWait(driver, 30).until(EC.visibility_of_element_located((By.XPATH, "//a[normalize-space(text())='Schedule Work Order Import']"))).click()
            time.sleep(2)
            ####Import Type Dropdown####
            WebDriverWait(driver, 30).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="Import_type"]')))
            dropdown_element = driver.find_element(By.XPATH, '//*[@id="Import_type"]')
            select = Select(dropdown_element)
            select.select_by_value("2")
            time.sleep(6)
            ####Import Template Dropdown####
            WebDriverWait(driver, 50).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="templateid"]')))
            dropdown_element = driver.find_element(By.XPATH, '//*[@id="templateid"]')
            select = Select(dropdown_element)
            select.select_by_value("14")
            time.sleep(1)
            # try:
            #     import_type = WebDriverWait(driver, 30).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="Import_type"]')))
            #     print("Import type found directly.")
            # except:
            #     print("Import type not found.")    
            #     WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH, "//a[normalize-space(text())='Work Order']"))).click()
            #     time.sleep(2)
            #     WebDriverWait(driver, 30).until(EC.element_to_be_clickable((By.XPATH, "//a[normalize-space(text())='Schedule Work Order Import']"))).click()
            #     time.sleep(2)
            #     import_type = WebDriverWait(driver, 30).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="Import_type"]')))
            # # ----- Import Type Dropdown -----
            # Select(import_type).select_by_value("2")
            # time.sleep(4)
            # # ----- Import Template Dropdown -----
            # template = WebDriverWait(driver, 30).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="templateid"]')))
            # Select(template).select_by_value("14")
            # time.sleep(4)
            ####Upload File####
            upload = WebDriverWait(driver, 30).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="fileToUpload"]')))
            upload.send_keys(raw_excel)
            print(f"Uploaded file: {raw_excel}")
            log_message(log_file, f"Uploaded file: {raw_excel}")
            WebDriverWait(driver, 30).until(EC.visibility_of_element_located((By.XPATH, "//button[normalize-space(text())='Schedule Task']")))
            schedule_task = driver.find_element(By.XPATH, "//button[normalize-space(text())='Schedule Task']")
            schedule_task.click()
            print(f"\033[92mUploaded {raw_excel} Work Order Import successfully.\033[0m")
            time.sleep(2)
            #### Mail Logic ####
            print("It's time to send a mail")
            recipient_emails = credentials['sdavis']
            cc_emails = credentials['ybotID']
            subject = f"Fuse Upload complete for {department}"
            body_message = f"The output file for the {department} department has been successfully uploaded into Fuse."
            body_message1 = ""
            attachment_path = ""
            try:
                Sendmail(recipient_emails, cc_emails, subject, body_message, body_message1, attachment_path)
                log_message(log_file, "Mail sent Successfully for uploaded Work Order bulk file.")
            except Exception:
                log_message(log_file, "Mail not sent for uploaded Work Order bulk file.")
            time.sleep(5)
             ###Import Status Table####
            print("Going to get file name uploaded rows")
            WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH,  "//table[@id='exampleImport']//tbody/tr[1]")))
            schedule_date_row = driver.find_element(By.XPATH, "//table[@id='exampleImport']//tbody/tr[1]")
            schedule_date = schedule_date_row.find_element(By.XPATH, "(//table[@id='exampleImport']//tbody/tr/td[4])[1]")
            date_str_1 = schedule_date.text.strip()
            print(f"Upload Date:{date_str_1}")
            upload_date = datetime.strptime(date_str_1, "%Y-%m-%d")
            print("Parsed date:", upload_date)
            formatted_date = upload_date.strftime("%d/%m/%Y")
            print("Formatted date:", formatted_date)
            today_date = datetime.now().strftime("%d/%m/%Y")
            print("Today date:", today_date)
            print(f"Schedule Date:{date_str_1}")
            log_message(log_file, f"Schedule Date:{date_str_1}")
            if today_date == formatted_date:
                WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH,  "//table[@id='exampleImport']//tr[@role='row'][1]/td[last()]/a")))
                file_name = driver.find_element(By.XPATH, "//table[@id='exampleImport']//tr[@role='row'][1]/td[last()]/a")
                get_file_name = file_name.text.strip()
                print(f"File Name:{get_file_name}")
                log_message(log_file, f"File Name:{get_file_name}")
                alter_file_name = get_file_name.split(".")[0]
                print(f"Alter File Name: {alter_file_name}")
                if department in alter_file_name:
                    print(f"\033[92mUpload verified for {department} department.\033[0m")
                    print("Department Name and File Name is matched.")
                    log_message(log_file, "Department Name and File Name is matched.")
                    update_completed_status(folder_path, department, "file uploaded")
                    
        except TimeoutException:
            print("\033[91mUpload failed due to timeout:\033[0m")
            log_message(log_file, "Upload failed due to timeout")
            print("It's time to send a mail")
            recipient_emails = credentials['ybotID']
            cc_emails = credentials['ybotID']
            subject = f"Fuse Upload failed for {department}"
            body_message = f"The output file for the {department} department failed to upload into Fuse."
            body_message1 = ""
            attachment_path = ""
            try:
                Sendmail(recipient_emails, cc_emails, subject, body_message, body_message1, attachment_path)
                log_message(log_file, "Mail sent Successfully for failed to upload Work Order bulk file.")
            except Exception:
                log_message(log_file, "Mail not sent for failed to upload Work Order bulk file.")
        



