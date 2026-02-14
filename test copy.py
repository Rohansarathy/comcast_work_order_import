import os
import json
import re
import time
import psutil
from selenium import webdriver
from datetime import datetime
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support import expected_conditions as EC


from Sendmail import Sendmail
# from fuselogin import fuse_login
from openpyxl import load_workbook
from bulk_upload import upload_raw_files
from edit_columns import insert_columns_from_main
from excel_utils import update_completed_status

credentials_file = 'accpet_loadjob.json'
with open(credentials_file, 'r') as file:
    credentials = json.load(file)

browsers = ['chrome.exe', 'firefox.exe', 'msedge.exe', 'chromedriver.exe']

# Closing all exisitng browser
for proc in psutil.process_iter(['name']):
    if proc.info['name'] in browsers:
        proc.kill()

def log_message(log_file_path, message):
    os.makedirs(os.path.dirname(log_file_path), exist_ok=True)
    with open(log_file_path, 'a', encoding='utf-8') as log:
        log.write(message + '\n')
    print(message)
log_file = f"{credentials['Logfile']}\\work_order_imprt_logs.txt"

chrome_options = Options()
chrome_options.add_argument("--log-level=3")
chrome_options.add_argument(r"--user-data-dir=C:\ChromeSeleniumProfiles\workorder")
chrome_options.add_experimental_option("detach", True)

service = Service(ChromeDriverManager("142.0.7444.60").install())
driver = webdriver.Chrome(service=service, options=chrome_options)
driver.maximize_window()
driver.execute_cdp_cmd("Page.bringToFront", {})

# FuseLogin = fuse_login(driver, credentials, log_file)
base_folder  = r"C:\Users\RohansarathyGoudhama\Downloads\work_order_import"
# base_folder = r"C:\Users\RohansarathyGoudhama\Downloads\Output"
folder_path = credentials['Main_file']
extaction_file = credentials['extaction_file']

# Mapping RAW filename → Lookup sheet name
key_to_lookup = {
    "mia1": "Miami.Broward VLookup",
    "pit1": "Keystone",
    "flor": "FTM V-Lookup",
    "har1": "WNE V-lookup",
    "ne01": "GBR V-Lookup",
    "bigs1": "Big South V Lookup",
    "bigs2": "Big South V Lookup",
    "atl1": "Big South V Lookup",
    "chi2": "Big South V Lookup",
    "denv": "Colorado V Lookup",
    "sea1": "Seattle V Lookup",
}

# departments = [
#     "mia1","jax1","belt","bay1","free","pit1","port","flor",
#     "har1","ne01","hou1","bigs1","atl1","bigs2","denv",
#     "salt","sea1","msp1","chi2"
# ]


departments = [
    "mia1" 
]
today = datetime.today().strftime("%m-%d-%Y") 
def ensure_remarks_header(folder_path):
    if not folder_path or not folder_path.lower().endswith('.xlsx') or not os.path.exists(folder_path):
        return None

    wb = load_workbook(folder_path)
    ws = wb.active

    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    if "Remarks" not in headers:
        ws.cell(row=1, column=len(headers) + 1, value="Remarks")
        remarks_col_idx = len(headers) + 1
        try:
            wb.save(folder_path)
            print(f"Added 'Remarks' header to: {folder_path}")
        except PermissionError:
            print(f"Permission denied when saving: {folder_path}. Close the file and retry.")
        finally:
            wb.close()
        return remarks_col_idx
    else:
        remarks_col_idx = headers.index("Remarks") + 1
        wb.close()
        return remarks_col_idx


def process_raw_file(folder_path, raw_excel, date_folder, extaction_file, department):
    print("PROCESSING RAW FILE:", raw_excel)
    # 1) First INSERT REQUIRED COLUMNS into RAW file
    print("STEP 1: Checking & inserting missing columns from MAIN\n")
    area_column, rf_column, upload_direct = insert_columns_from_main(folder_path, raw_excel, department)
    if upload_direct:
        print("Upload Direct Raw files.\n")
        upload_raw_files(driver, folder_path, date_folder, credentials, raw_excel, department, log_file)
        return
    # else:
    if not (area_column or rf_column):
        print("No new columns inserted. Columns already existed.\n")
    print(f"Department detected from folder: {department}")

    if department not in key_to_lookup:
        print(f"No lookup mapping found for department '{department}'\n")
        return

    lookup_sheet = key_to_lookup[department]
    print(f"RAW - Lookup sheet selected: {lookup_sheet}")
    raw_wb = load_workbook(raw_excel)
    raw_ws = raw_wb.active
    headers = [str(c.value).strip() if c.value else "" for c in raw_ws[1]]

    if "RteC" not in headers:
        print("RteC column missing in RAW file.\n")
        return

    rtec_col = headers.index("RteC") + 1

    # Read RteC values
    rte_values = []
    for row in range(2, raw_ws.max_row + 1):
        cell = raw_ws.cell(row=row, column=rtec_col)
        value = cell.value

        if value in (None, ""):
            continue

        value_str = str(value).strip()

        # Keep EVERYTHING as string
        cell.value = value_str
    print(f"Total RteC values found: {len(rte_values)}")

    lookup_wb = load_workbook(extaction_file)

    if lookup_sheet not in lookup_wb.sheetnames:
        print(f"Lookup sheet '{lookup_sheet}' not in extraction file.\n")
        return

    lookup_ws = lookup_wb[lookup_sheet]
    matched_rows = []
    for row in lookup_ws.iter_rows(min_row=2, values_only=True):
        if row and row[0]:
            lookup_key = str(row[0]).strip()
            if lookup_key in rte_values:
                clean = [v for v in row if v is not None]
                matched_rows.append(clean)


    print(f"\nMatched {len(matched_rows)} rows:")
    for r in matched_rows:
        print("--->", r)

    if not matched_rows:
        print("No matching RteC values.\n")
        return

    # Target Column
    headers_lower = [h.lower() for h in headers]

    if "area" in headers_lower:
        target_col = headers_lower.index("area") + 1
    elif "rf" in headers_lower:
        target_col = headers_lower.index("rf") + 1

    else:
        print("\033[91mError: AREA/RF column still not found after insertion.\033[0m")
        recipient_emails = credentials['ybotID']
        cc_emails = credentials['ybotID']
        subject = f" AREA/RF column still not found for {department}"
        body_message = f"AREA/RF column still not found  for {department}. Kindly check the issue."
        body_message1 = ""
        attachment_path = ""
        try:
            Sendmail(recipient_emails, cc_emails, subject, body_message, body_message1, attachment_path)
            log_message(log_file, "Mail sent Successfully for AREA/RF column still not found.")
        except Exception:
            log_message(log_file, "Mail not sent for AREA/RF column still not found.")
        return

    # lookup_dict = {row[0]: row[1] for row in matched_rows}
    lookup_dict = {str(row[0]).strip(): row[1] for row in matched_rows}

    update_count = 0
    for row in range(2, raw_ws.max_row + 1):
        rtec_value = raw_ws.cell(row=row, column=rtec_col).value
        if not rtec_value:
            continue

        rtec_value = str(rtec_value).strip()
        if rtec_value in lookup_dict:
            raw_ws.cell(row=row, column=target_col).value = lookup_dict[rtec_value]
            update_count += 1

    raw_wb.save(raw_excel)
    print(f"\n Updated {update_count} rows successfully.")
    print("Uploading file after lookup update...")
    upload_raw_files(driver, folder_path, date_folder, credentials, raw_excel, department, log_file)
    print("*****************************************************\n")


folder_path = credentials['Main_file'] 
dynamic_wb = load_workbook(folder_path) 
dynamic_ws = dynamic_wb.active
dept_status = {}
headers = [str(c.value).strip() if c.value else "" for c in dynamic_ws[1]]
status_col = headers.index("Status") + 1
dept_col = headers.index("Department") + 1

for row in range(2, dynamic_ws.max_row + 1):
    dept = dynamic_ws.cell(row=row, column=dept_col).value
    status = dynamic_ws.cell(row=row, column=status_col).value
    if dept:
        dept_status[str(dept).strip().lower()] = str(status).strip() if status else ""
for dept in departments:
    print("\n======================================")
    print(f"Checking Department: {dept}")
    if dept in dept_status and dept_status[dept].lower() == "no job found":  
        update_completed_status(folder_path, dept, "No Job Found")
        print(f"\033[93mSkipping department '{dept}' because Status = 'No Job Found'\033[0m")
        recipient_emails = credentials['ybotID']
        cc_emails = credentials['ybotID']
        subject = f" {departments} Department No Job Found"
        body_message = f"Skipping department '{dept}' because Status = 'No Job Found'"
        body_message1 = ""
        attachment_path = ""
        try:
            Sendmail(recipient_emails, cc_emails, subject, body_message, body_message1, attachment_path)
            log_message(log_file, "Mail sent Successfully for Department folder not found.")
        except Exception:
            log_message(log_file, "Mail not sent for Department folder not found.")
        continue

    dept_folder = os.path.join(base_folder, dept)
    if not os.path.isdir(dept_folder):
        print(f"Department folder not found: {dept_folder}")
        # recipient_emails = credentials['ybotID']
        # cc_emails = credentials['ybotID']
        # subject = f" {departments} Department folder not found"
        # body_message = f"{departments} Department folder not found.Kindly check the issue."
        # body_message1 = ""
        # attachment_path = ""
        # try:
        #     Sendmail(recipient_emails, cc_emails, subject, body_message, body_message1, attachment_path)
        #     log_message(log_file, "Mail sent Successfully for Department folder not found.")
        # except Exception:
        #     log_message(log_file, "Mail not sent for Department folder not found.")
        continue

    date_folder = os.path.join(dept_folder, today)

    if not os.path.isdir(date_folder):
        print(f"Date folder not found: {date_folder}")
        continue

    excel_files = [
        f for f in os.listdir(date_folder)
        if f.lower().endswith(".xlsx")
    ]

    if not excel_files:
        print("No Excel file found in date folder")
        recipient_emails = credentials['ybotID']
        cc_emails = credentials['ybotID']
        subject = f" {departments} Excel file not found"
        body_message = f"{departments} Department Excel file not found."
        body_message1 = ""
        attachment_path = ""
        try:
            Sendmail(recipient_emails, cc_emails, subject, body_message, body_message1, attachment_path)
            log_message(log_file, "Mail sent Successfully for Department Excel file not found.")
        except Exception:
            log_message(log_file, "Mail not sent for Department Excel file not found.")
        continue

    for excel_name in excel_files:
        excel_path  = os.path.join(date_folder, excel_name)
        print(f"Processing file: {excel_path}")

        process_raw_file(folder_path, excel_path, date_folder, extaction_file, dept)



